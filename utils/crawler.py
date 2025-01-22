# utils/crawler.py

import os
import time
import random
import datetime
import openpyxl
import re
import shutil  # 用于copyfile
import threading
import requests

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service as EdgeService
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from concurrent.futures import ThreadPoolExecutor, as_completed

MAX_WORKERS = 3

def progress_log_callback(callback, message):
    """统一日志输出"""
    if callback:
        callback(message)
    else:
        print(message)

def random_delay(min_sec=0.5, max_sec=1.5):
    """避免爬取过快"""
    time.sleep(random.uniform(min_sec, max_sec))

def safe_execute(func):
    """装饰器统一处理异常"""
    def wrapper(*args, progress_callback=None, **kwargs):
        try:
            return func(*args, progress_callback=progress_callback, **kwargs)
        except Exception as e:
            progress_log_callback(progress_callback, f"执行出错: {str(e)}")
            if 'progress_percent_callback' in kwargs:
                kwargs['progress_percent_callback'](100, 0)
            return None
    return wrapper

# -----------------------------------------------------------------------------
# 新游爬虫
# -----------------------------------------------------------------------------
@safe_execute
def crawl_new_games(
    start_date_str,
    end_date_str,
    progress_callback=None,
    enable_version_match=True,
    progress_percent_callback=None
):
    """
    1. 无"序号"列；列头: [ "日期", "游戏名称", "状态", "厂商", "类型", "评分" ]
    2. 分天并发，每日爬完就输出并写到 Excel，但最终爬完后，再做一次"按日期升序"全表排序。
    3. 若 enable_version_match=True，则自动在同一个 Excel 里匹配版号并存储（不改名/不另存）。
    """

    # 提示
    progress_log_callback(progress_callback,
        "任务开始，可能存在网络缓慢等情况，请耐心等待。")

    try:
        sd = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
        ed = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
        if ed < sd:
            msg="结束日期不能早于起始日期，操作已取消。"
            progress_log_callback(progress_callback, msg)
            if progress_percent_callback:
                progress_percent_callback(100, 0)
            return
    except:
        msg="日期格式不正确，请使用 yyyy-MM-dd。操作已取消。"
        progress_log_callback(progress_callback, msg)
        if progress_percent_callback:
            progress_percent_callback(100, 0)
        return

    date_list=[]
    delta=(ed - sd).days+1
    for i in range(delta):
        date_list.append(sd + datetime.timedelta(days=i))

    total_dates=len(date_list)
    if total_dates==0:
        progress_log_callback(progress_callback,
            "无可爬取日期。")
        if progress_percent_callback:
            progress_percent_callback(100,0)
        return

    # 创建Excel
    today_str = datetime.date.today().strftime("%Y-%m-%d")
    excel_filename = f"游戏数据_{today_str}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title="游戏数据"
    ws.append(["日期","游戏名称","状态","厂商","类型","评分"])  # 无"序号"
    wb.save(excel_filename)

    progress_log_callback(progress_callback,
        f"共需爬取 {total_dates} 天的新游信息，将分天爬取...")

    def crawl_one_day(d):
        """
        返回 (day_str, [ (name, status, man, types, rating) ])，若结构异常 => raise RuntimeError
        """
        day_str = d.strftime("%Y-%m-%d")
        opt = webdriver.EdgeOptions()
        driver = webdriver.Edge(
            service=EdgeService(EdgeChromiumDriverManager().install()),
            options=opt
        )
        results = []
        try:
            url = f"https://www.taptap.cn/app-calendar/{day_str}"
            driver.get(url)
            
            # 修改：先等待页面加载完成
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 
                        "div.daily-event-list__content"))
                )
            except:
                raise RuntimeError("网页结构疑似存在更新变动，请联系开发者进行解决。")

            # 查找游戏元素
            games = driver.find_elements(
                By.CSS_SELECTOR,
                "div.daily-event-list__content > a.tap-router"
            )
            
            # 如果没有游戏，直接返回空列表
            if not games:
                return (day_str, [])

            for g in games:
                random_delay()
                try:
                    name_el=g.find_element(
                        By.CSS_SELECTOR,"div.daily-event-app-info__title")
                    name=name_el.get_attribute("content").strip()
                except:
                    name="未知名称"
                try:
                    t_list=g.find_elements(
                        By.CSS_SELECTOR,
                        "div.daily-event-app-info__tag div.tap-label-tag")
                    types="/".join([t.text.strip() for t in t_list])
                except:
                    types="未知类型"
                try:
                    rating_el=g.find_element(
                        By.CSS_SELECTOR,
                        "div.daily-event-app-info__rating .tap-rating__number")
                    rating=rating_el.text.strip()
                except:
                    rating="未知评分"
                try:
                    st_el=g.find_element(
                        By.CSS_SELECTOR,"span.event-type-label__title")
                    status=st_el.text.strip()
                except:
                    try:
                        st2=g.find_element(
                            By.CSS_SELECTOR,"div.event-recommend-label__title")
                        status=st2.text.strip()
                    except:
                        status="未知状态"

                # 子页面
                href=g.get_attribute("href")
                if not href.startswith("http"):
                    href="https://www.taptap.cn"+href
                driver.execute_script("window.open(arguments[0]);", href)
                driver.switch_to.window(driver.window_handles[1])
                random_delay()

                man="未知厂商"
                try:
                    WebDriverWait(driver,5).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR,
                            "div.single-info__content"))
                    )
                    blocks=driver.find_elements(By.CSS_SELECTOR,"div.single-info__content")
                    priority=["厂商","发行","发行商","开发"]
                    info_dict={}
                    for b in blocks:
                        try:
                            lab=b.find_element(
                                By.CSS_SELECTOR,".caption-m12-w12").text.strip()
                            val=b.find_element(
                                By.CSS_SELECTOR,".single-info__content__value").text.strip()
                            info_dict[lab]=val
                        except:
                            pass
                    for pk in priority:
                        if pk in info_dict:
                            man=info_dict[pk]
                            break
                except:
                    pass

                driver.close()
                driver.switch_to.window(driver.window_handles[0])

                results.append( (name, status, man, types, rating) )
        finally:
            driver.quit()
        return (day_str, results)

    completed=0

    # 并发: 以天为粒度
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_map={ executor.submit(crawl_one_day, d): d for d in date_list}
        for future in as_completed(future_map):
            d = future_map[future]
            try:
                day_str, day_data = future.result()
            except Exception as e:
                progress_log_callback(progress_callback, str(e))
                if progress_percent_callback:
                    progress_percent_callback(100,0)
                return

            if day_data:
                wb_cur=openpyxl.load_workbook(excel_filename)
                ws_cur=wb_cur.active
                block=[f"\n=== [日期 {day_str}, 共{len(day_data)} 款游戏] ==="]
                for (nm, st, man, ty, rt) in day_data:
                    ws_cur.append([day_str,nm,st,man,ty,rt])
                    block.append(
                        f"  * {nm}\n"
                        f"    状态：{st}\n"
                        f"    厂商：{man}\n"
                        f"    类型：{ty}\n"
                        f"    评分：{rt}"
                    )
                wb_cur.save(excel_filename)
                text="\n".join(block)
                progress_log_callback(progress_callback, text)
            else:
                progress_log_callback(progress_callback,
                    f"=== [日期 {day_str}] 无游戏信息 ===")

            completed+=1
            if progress_percent_callback:
                val=int(completed*100/total_dates)
                progress_percent_callback(val,0)

    # 全部爬完后 => 对整个Excel按"日期"升序排序
    final_wb=openpyxl.load_workbook(excel_filename)
    final_ws=final_wb.active
    data_rows=list(final_ws.values)
    headers=data_rows[0]
    body=data_rows[1:]
    def parse_date(ds):
        try:
            return datetime.datetime.strptime(ds, "%Y-%m-%d")
        except:
            return datetime.datetime(1970,1,1)
    body.sort(key=lambda row: parse_date(row[0]))
    final_ws.delete_rows(1, final_ws.max_row)
    final_ws.append(headers)
    for row in body:
        final_ws.append(row)
    final_wb.save(excel_filename)

    progress_log_callback(progress_callback,
        f"新游数据已保存至 {excel_filename} (已按日期升序整理)")

    if progress_percent_callback:
        progress_percent_callback(100,0)

    # 若自动版号匹配
    if enable_version_match:
        if progress_percent_callback:
            progress_percent_callback(0,1)
        match_version_numbers(
            excel_filename,
            progress_callback=progress_callback,
            progress_percent_callback=progress_percent_callback,
            stage=1,
            create_new_file=False
        )

# -----------------------------------------------------------------------------
# 版号匹配
# -----------------------------------------------------------------------------
def match_version_numbers(
    excel_filename,
    progress_callback=None,
    progress_percent_callback=None,
    stage=1,
    create_new_file=True
):
    """
    - 如果 create_new_file=True => 基于原文件创建副本，并在副本上进行后续操作
    - 如果 create_new_file=False => 在同文件追加
    - 分段输出(每2~3行)
    - 需添加表头: [ "游戏名称", "出版单位", "运营单位", "文号", "出版物号", "时间", "游戏类型", "申报类别", "是否多个结果" ]
    """
    import shutil  # 用于复制文件

    progress_log_callback(progress_callback,
        "开始版号匹配，可能存在网络缓慢等情况，请耐心等待。")

    if not os.path.exists(excel_filename):
        progress_log_callback(progress_callback,
            f"文件 {excel_filename} 不存在，无法进行版号匹配。")
        if progress_percent_callback:
            progress_percent_callback(100, stage)
        return

    # 若要求另存为新文件，则先创建副本，并使用该副本进行操作
    if create_new_file:
        base, ext = os.path.splitext(excel_filename)
        new_fname = f"{base}-已匹配版号{ext}"
        try:
            shutil.copyfile(excel_filename, new_fname)
            progress_log_callback(progress_callback,
                f"已基于原文件创建副本 {new_fname}，开始在副本上匹配版号。")
            excel_filename = new_fname  # 后续操作使用新副本
        except Exception as copy_err:
            progress_log_callback(progress_callback,
                f"创建文件副本失败: {copy_err}")
            if progress_percent_callback:
                progress_percent_callback(100, stage)
            return

    try:
        wb = openpyxl.load_workbook(excel_filename)
    except:
        progress_log_callback(progress_callback,
            f"无法加载Excel：{excel_filename}")
        if progress_percent_callback:
            progress_percent_callback(100, stage)
        return

    ws = wb.active
    new_headers = [
        "游戏名称","出版单位","运营单位","文号","出版物号",
        "时间","游戏类型","申报类别","是否多个结果"
    ]
    current_cols = [c.value for c in ws[1]]
    for nh in new_headers:
        if nh not in current_cols:
            ws.cell(row=1, column=len(current_cols)+1, value=nh)
            current_cols.append(nh)
    wb.save(excel_filename)

    name_col = None
    for idx, c in enumerate(current_cols, start=1):
        if c in ["游戏名称","名称"]:
            name_col = idx
            break

    if name_col is None:
        progress_log_callback(progress_callback,
            "未找到'游戏名称'或'名称'列，跳过版号匹配。")
        if progress_percent_callback:
            progress_percent_callback(100, stage)
        return

    game_list = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        val = row[name_col - 1].value
        if val:
            game_list.append((row_idx, val))

    total_count = len(game_list)
    if total_count == 0:
        progress_log_callback(progress_callback,
            "没有需要匹配版号的条目。")
        if progress_percent_callback:
            progress_percent_callback(100, stage)
        return

    progress_log_callback(progress_callback,
        f"需要匹配 {total_count} 条游戏数据...")

    #======== 修复点：写回9列时，使用表头定位列索引 ========
    fields_for_match = [
        "游戏名称",   # info[0]
        "出版单位",   # info[1]
        "运营单位",   # info[2]
        "文号",       # info[3]
        "出版物号",   # info[4]
        "时间",       # info[5]
        "游戏类型",   # info[6]
        "申报类别",   # info[7]
        "是否多个结果" # info[8]
    ]

    cache = {}
    def fetch_game_info(g_name):
        if g_name in cache:
            return cache[g_name]
        opt = webdriver.EdgeOptions()
        driver = webdriver.Edge(
            service=EdgeService(EdgeChromiumDriverManager().install()),
            options=opt
        )
        res = None
        try:
            qn = re.sub(r'（[^）]*）', '', g_name)
            qn = re.sub(r'\([^)]*\)', '', qn).strip()
            url = (f"https://www.nppa.gov.cn/bsfw/jggs/cxjg/index.html?"
                   f"mc={qn}&cbdw=&yydw=&wh=undefined&description=#")
            driver.get(url)
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "#dataCenter"))
                )
            except:
                raise RuntimeError("网页结构疑似存在更新变动，请联系开发者进行解决。")

            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#dataCenter tr"))
                )
            except:
                rows2 = []
            else:
                rows2 = driver.find_elements(By.CSS_SELECTOR, "#dataCenter tr")

            if not rows2:
                res = None
            else:
                multiple_flag = "是" if len(rows2) > 1 else "否"
                exact_elem = None
                for rr in rows2:
                    time.sleep(0.3)
                    try:
                        t = rr.find_element(By.CSS_SELECTOR, "a").text.strip()
                        if t == g_name:
                            exact_elem = rr
                            break
                    except:
                        pass
                if not exact_elem:
                    exact_elem = rows2[0]
                res = extract_game_info(exact_elem, multiple_flag, driver)
        finally:
            driver.quit()

        cache[g_name] = res
        return res

    def extract_game_info(elem, multi_flag, driver):
        try:
            tds = elem.find_elements(By.TAG_NAME, "td")
            if len(tds) < 7:
                return None
            gn = tds[1].text.strip()
            pub = tds[2].text.strip()
            op = tds[3].text.strip()
            appr = tds[4].text.strip()
            pubn = tds[5].text.strip()
            ds = tds[6].text.strip()

            detail_url = ""
            try:
                detail_url = tds[1].find_element(By.TAG_NAME, "a").get_attribute("href")
            except:
                pass

            gtype, appcat = "", ""
            if detail_url:
                driver.execute_script("window.open(arguments[0]);", detail_url)
                driver.switch_to.window(driver.window_handles[1])
                random_delay()
                lines = driver.find_elements(By.CSS_SELECTOR, ".cFrame.nFrame table tr")
                for ln in lines:
                    try:
                        lab = ln.find_element(By.XPATH, "./td[1]").text.strip()
                        val = ln.find_element(By.XPATH, "./td[2]").text.strip()
                        if lab == "游戏类型":
                            gtype = val
                        elif lab == "申报类别":
                            appcat = val
                    except:
                        pass
                driver.close()
                driver.switch_to.window(driver.window_handles[0])

            return [gn, pub, op, appr, pubn, ds, gtype, appcat, multi_flag]
        except:
            return None

    partial_flush_size = 3
    completed = 0
    buffered = []
    results_map = {}

    def flush_results(buf):
        if not buf:
            return []
        buf.sort(key=lambda x: x[0])  # row_idx升序
        lines = []
        wb2 = openpyxl.load_workbook(excel_filename)
        ws2 = wb2.active
        # 取一次完整的表头(字符串列表)
        head_cols = [c.value for c in ws2[1]]

        for (r_idx, g_n, info) in buf:
            if info:
                # info顺序: [gn, pub, op, appr, pubn, ds, gtype, appcat, multi_flag]
                # 逐一写入表头对应的列
                block = [
                    f"--- [行 {r_idx}, 游戏名: {g_n}] ---",
                    f"  游戏名称：{info[0]}",
                    f"  出版单位：{info[1]}",
                    f"  运营单位：{info[2]}",
                    f"  文号：{info[3]}",
                    f"  出版物号：{info[4]}",
                    f"  时间：{info[5]}",
                    f"  游戏类型：{info[6]}",
                    f"  申报类别：{info[7]}",
                    f"  是否多个结果：{info[8]}"
                ]
                lines.extend(block)
                # 按表头名字依次写入
                for i, field_name in enumerate([
                    "游戏名称","出版单位","运营单位","文号","出版物号",
                    "时间","游戏类型","申报类别","是否多个结果"
                ]):
                    # 找到该列
                    if field_name in head_cols:
                        col_idx = head_cols.index(field_name) + 1  # 1-based
                        ws2.cell(row=r_idx, column=col_idx, value=info[i])
                    # 若找不到就不写
            else:
                lines.append(f"--- [行 {r_idx}, 游戏名: {g_n}] => 未查询到该游戏版号信息。")

        wb2.save(excel_filename)
        text = "\n".join(lines)
        progress_log_callback(progress_callback, text)
        return []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        future_map = {}
        for (r_idx, g_n) in game_list:
            future = ex.submit(fetch_game_info, g_n)
            future_map[future] = (r_idx, g_n)

        for future in as_completed(future_map):
            row_i, g_na = future_map[future]
            try:
                info = future.result()
            except Exception as e:
                progress_log_callback(progress_callback, str(e))
                if progress_percent_callback:
                    progress_percent_callback(100, stage)
                return
            results_map[row_i] = info
            buffered.append((row_i, g_na, info))
            completed += 1
            if len(buffered) >= partial_flush_size:
                buffered = flush_results(buffered)
            if progress_percent_callback:
                pr = int(completed * 100 / total_count)
                progress_percent_callback(pr, stage)

    if buffered:
        buffered = flush_results(buffered)

    # 若 create_new_file=True，已在开头复制并操作副本，不再另存
    if create_new_file:
        pass
    else:
        progress_log_callback(progress_callback,
            f"版号匹配完成，结果已追加到原文件 {excel_filename}")

    if progress_percent_callback:
        progress_percent_callback(100, stage)
