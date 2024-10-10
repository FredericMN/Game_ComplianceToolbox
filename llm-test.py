# analyze_excel.py

import os
import sys
from utils.large_model import (
    check_model_configured, check_and_download_model, load_classifier
)
from openpyxl import load_workbook

def main():
    # 指定需要检测的Excel文件路径
    excel_file_path = r"D:\AI\模型测试3.xlsx"

    # 检查文件是否存在
    if not os.path.exists(excel_file_path):
        print(f"文件不存在: {excel_file_path}")
        sys.exit(1)

    # 确保模型已配置，如果未配置则下载
    try:
        if not check_model_configured():
            print("大模型未配置，正在下载...")
            check_and_download_model(progress_callback=print)
    except Exception as e:
        print(f"下载大模型时发生错误: {e}")
        sys.exit(1)

    # 加载分类器
    try:
        classifier = load_classifier(device='cpu')
    except Exception as e:
        print(f"加载分类器时发生错误: {e}")
        sys.exit(1)

    # 加载Excel工作簿
    try:
        wb = load_workbook(excel_file_path)
    except Exception as e:
        print(f"加载Excel文件时发生错误: {e}")
        sys.exit(1)

    # 处理每个工作表
    for sheet in wb.worksheets:
        print(f"正在处理工作表: {sheet.title}")
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                if cell.value:
                    text = str(cell.value).strip()
                    if text:
                        try:
                            # 使用分类器分析文本
                            results = classifier(text, top_k=None, truncation=True)
                            # 将结果转换为标签及其分数的字典
                            label_scores = {item['label']: item['score'] for item in results}
                            
                            # 打印分析结果
                            print(f"单元格 {cell.coordinate} 内容: {text}")
                            for label, score in label_scores.items():
                                print(f"  {label}: {score:.2f}")
                            print()

                            # 将标签分数附加到单元格内容后
                            appended_text = f"{text} ["
                            appended_text += ", ".join([f"{label}:{score:.2f}" for label, score in label_scores.items()])
                            appended_text += "]"
                            cell.value = appended_text

                        except Exception as e:
                            print(f"分析单元格 {cell.coordinate} 时发生错误: {e}")

    # 保存分析后的Excel文件
    new_file_path = excel_file_path.replace('.xlsx', '_analyzed.xlsx')
    try:
        wb.save(new_file_path)
        print(f"分析完成，结果已保存为: {new_file_path}")
    except Exception as e:
        print(f"保存分析结果时发生错误: {e}")

if __name__ == "__main__":
    main()
